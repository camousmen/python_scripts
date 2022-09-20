from calendar import month
import json
import requests
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

MONTH_RU = {
    'January':'Январь',
    'February':'Февраль',
    'March':'Март',
    'April':'Апрель',
    'May':'Май',
    'June':'Июнь',
    'July':'Июль',
    'August':'Август',
    'September':'Сентябрь',
    'October':'Октябрь',
    'November':'Ноябрь',
    'December':'Декабрь'
}

STORE_EXCEL_DICT = {
    'Магазин "Мааппа"': 'маг1_1с',
    'Магазин "Универсам"': 'маг2_1с',
    'Магазин №5': 'маг5_1с',
    'Магазин №8': 'маг8_1с',
    'Магазин №10': 'маг10_1с'
}


def date_to_russia(date_string):
    """Перевод даты на русский"""
    date_string = date_string.split()
    date_string[0] = MONTH_RU[date_string[0]]
    date_string = str(date_string[0]) + ' ' + str(date_string[1])
    return date_string

def get_store_list():
    """Список магазинов"""
    url = f'http://Server/Retail/odata/standard.odata/Catalog_Магазины?$format=json'
    payload = {}
    headers = {'Authorization': 'Basic b2RhdGEuY2Ftb3VzbWVuOklkZWEyMDIy'}
    
    response = requests.request("GET", url, headers=headers, data=payload)
    value_json = json.loads(response.text)['value']

    store_dict = {}
    for el in value_json:
        store_dict[el['Ref_Key']] = el['Description']
    return store_dict


def get_orp_on_date(date_string):
    """Возвращает ОРП на указанную дату"""
    orp_list = []
    start_time = f'{date_string}T00:00:00'
    end_time = f'{date_string}T23:59:59'
    
    url = (f'http://Server/Retail/odata/standard.odata/Document_ОтчетОРозничныхПродажах?'
           f'$filter=Date%20ge%20datetime%27{start_time}%27%20and%20Date%20le%20datetime%27{end_time}%27&$format=json')
    payload = {}
    headers = {'Authorization': 'Basic b2RhdGEuY2Ftb3VzbWVuOklkZWEyMDIy'}
    
    response = requests.request("GET", url, headers=headers, data=payload)
    value_json = json.loads(response.text)['value']

    for el in value_json:
        # сперва суммируем оплату картой
        card_pay = 0
        for p in el['ОплатаПлатежнымиКартами']:
            card_pay += p['Сумма']
        # информация о выручке
        orp_info = {
            'Дата': el['Date'],
            'Магазин': el['Магазин_Key'],
            'Наличные': el['СуммаОплатыНаличных'],
            'Терминал': card_pay,
            'Сертификат': el['ОплатаПодарочнымиСертификатами']
        }
        orp_list.append(orp_info)

    store_dict = get_store_list()
    stores_day_orp = {}
    for el in orp_list:
        stores_day_orp[el['Магазин']] = {
            'Дата':el['Дата'], 'Наличные':0, 'Терминал':0, 'Сертификат':0
        }

    for el in orp_list:
        stores_day_orp[el['Магазин']]['Наличные'] += el['Наличные']
        stores_day_orp[el['Магазин']]['Терминал'] += el['Терминал']
        stores_day_orp[el['Магазин']]['Сертификат'] += el['Сертификат']

    buf_value = {}
    for key, value in stores_day_orp.items():
        buf_value[store_dict[key]] = value

    return buf_value


cur_date = datetime.datetime.now()
cur_month = cur_date.strftime('%B %y')
cur_month = date_to_russia(cur_month)

current_day = datetime.date.today()
current_month = datetime.date.today().strftime('%Y-%m')

month_orp = []

cur_day_number = int(current_day.strftime('%d'))

for day in range(1, cur_day_number+1):
    if day < 10:
        day = f'0{str(day)}'
    else:
        day = str(day)
    day = f'{current_month}-{day}'
    month_orp.append(get_orp_on_date(day))


# запись полученных значений в эксель файл 
wb = load_workbook('test_excel.xlsx')

#В случае отсутствия листа нужно создать новый на основании шаблона
#if cur_month not in wb.sheetnames:
    #pass
sheet = wb[cur_month]

for day in range(1, len(month_orp)+1):
    for key, value in month_orp[day-1].items():
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == f'{STORE_EXCEL_DICT[key]}_{day}':
                    x = cell.row
                    y = cell.column
                    print(cell.row, cell.column)
                    # порядок записи - наличные, терминал, мобанк, сертификат
                    cell.value = value['Наличные']
                    cell.font = Font(color='008080')
                    sheet.cell(row=x+1, column=y).value = value['Терминал']
                    sheet.cell(row=x+1, column=y).font = Font(color='008080')
                    sheet.cell(row=x+3, column=y).value = value['Сертификат']
                    sheet.cell(row=x+3, column=y).font = Font(color='008080')

wb.save('text_output.xlsx')

